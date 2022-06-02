#! python3
# Sends emails based on payment status in spreadsheet.

# 1. Read data from an Excel spreadsheet.
# 2. Find all members who have not paid dues for the latest month.
# 3. Find their email addresses and send them personalized reminders.

# This means your code will need to do the following:

# 1. Open and read the cells of an Excel document with the openpyxl
# module. (See Chapter 13 for working with Excel files.)
# 2. Create a dictionary of members who are behind on their dues.

import openpyxl, smtplib, sys

# Oopen the spreadsheet and get the latest dues status.
wb = openpyxl.load_workbook('Reminders.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')

lastCol = sheet.max_column
latestMonth = sheet.cell(row=1, column=lastCol).value

# Check each memer's payment status
unpaidMembers = {}

for r in range(2, sheet.max_row + 1):
    payment = sheet.cell(row=r, column=lastCol).value
    if payment != 'paid':
        name = sheet.cell(row=r, column=1).value
        email = sheet.cell(row=r, column=2).value
        unpaidMembers[name] = email

# Log in to email account
smtpObj = smtplib.SMTP('smtp.example.com', 587)
smtpObj.ehlo()
smtpObj.starttls()
smtpObj.login('my_email_address@example.com', sys.argv[1])

# Send out reminder emails
for name, email in unpaidMembers.items():
    body = "Subject: %s dues unpaid.\nDear %s, \nRecords show that you hae not paid dues for %s. Please make this payment as soon as possible. Thank you!" % (latestMonth, name, latestMonth)
    print('Sending email to %s...' % email)
    sendmailStatus = smtpObj.sendmail('my_email_address@example.com', email, body)

    if sendmailStatus != {}:
        print('There was a problem sending email to %s: %s' % (email, sendmailStatus))

smtpObj.quit()