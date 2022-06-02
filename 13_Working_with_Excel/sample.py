import openpyxl

# Opening Excel Documents with OpenPyXL

# wb = openpyxl.load_workbook('example.xlsx')
# print(type(wb))

######################################################################

# Getting Sheets from the Workbook

# wb = openpyxl.load_workbook('wxample.xlsx')
# print(wb.sheetnames) # The workbook's sheets' names.

# sheet = wb['Sheet3'] # Get a sheet from the workbook.
# print(sheet)
# print(type(sheet))
# print(sheet.title) # Get the sheet's title as a string.
# anotherSheet = wb.ative # Get the active sheet.
# print(anotherSheet)

######################################################################

# Getting Cells from the Sheets

# wb = openpyxl.load_workbook('example.xlsx')
# sheet = wb['Sheet1'] # Get a sheet from the workbook.
# print(sheet['A1']) # Get a cell from the sheet.
# print(sheet['A1'].value)
# c = sheet['B1']
# print(c.value)
# print('Row %s, Column %s is %s' % (c.row, c.column, c.value)) # Get the row, column, and value from the cell.
# print('Cell %s is %s' % (c.coordinate, c.value))
# print(sheet['C1'].value)
# print(sheet.cell(row=1, column=2))
# print(sheet.cell(row=1, column=2).value)
# for i in range(1, 8, 2): # Go through every other row:
#     print(i, sheet.cell(row=i, column=2).value)


# wb = openpyxl.load_workbook('example.xlsx')
# sheet = wb['Sheet1']
# print(sheet.max_row) # Get the highest row number.
# print(sheet.max_column) # Get the highes column number.

######################################################################

# Converting Between Column Letters and Numbers

# from openpyxl.utils import get_column_letter, column_index_from_string
# print(get_column_letter(1)) # Translate column 1 to a letter.
# print(get_column_letter(2))
# print(get_column_letter(27))
# print(get_column_letter(900))

# wb = openpyxl.load_workbook('example.xlsx')
# sheet = wb['Sheet1']
# print(get_column_letter(sheet.max_column))
# print(column_index_from_string('A')) # Get A's number
# print(column_index_from_string('AA'))

######################################################################

# Getting Tows and Columns from the Sheets

# wb = openpyxl.load_workbook('example.xlsx')
# sheet = wb['Sheet1']
# print(tuple(sheet['A1' : 'C3'])) # Get all cells from A1 to C3.

# for rowOfCellObjects in sheet['A1' : 'C3']:
#     for cellObj in rowOfCellObjects:
#         print(cellObj.coordinate, cellObj.value)
#     print('--- END OF ROW ---')

######################################################################

# Workbooks, Sheets, Cells

# 1. Import the openpyxl module.
# 2. Call the openpyxl.load_workbook() function.
# 3. Get a Workbook object.
# 4. Use the active or sheetnames attributes.
# 5. Get a Worksheet object.
# 6. Use indexing or the cell() sheet method with row and column keyword
# arguments.
# 7. Get a Cell object.
# 8. Read the Cell object’s value attribute.

######################################################################

# Creating and Saving Excel Documents

# wb = openpyxl.Workbook() # Create a blank workbook.
# print(wb.sheetnames) # It starts with one sheets.
# sheet = wb.active
# print(sheet.title)
# sheet.title = 'Spam Bacon Eggs Sheet' # Change title
# print(wb.sheetnames)

# wb = openpyxl.load_workbook('example.xlsx')
# sheet = wb.active
# sheet.title = 'Spam Spam Spam'
# wb.save('example_copy.xlsx') # Save the workbook.

######################################################################

# Creating and Removing Sheets

# wb = openpyxl.Workbook()
# wb.sheetnames
# print(wb.create_sheet()) # Add a new sheet.
# print(wb.sheetnames)
# # Create a new sheet at index 0.
# print(wb.create_sheet(index=0, title='First Sheet'))
# print(wb.sheetnames)
# print(wb.create_sheet(index=2, title='Middle Sheet'))
# print(wb.sheetnames)

# print(wb.sheetnames)
# del wb['Middle Sheet']
# del wb['Sheet1']
# print(wb.sheetnames)

######################################################################

# Writing Values to Cells

# wb = openpyxl.Workbook()
# sheet = wb['Sheet']
# sheet[A1] = 'Hello, world!' # Edit the cell's value
# print(sheet['A1'].value)

######################################################################

# Setting the Font Style of Cells

# from openpyxl.style import Font
# wb = openpyxl.Workbook()
# sheet = wb['Sheet']
# italic24Font = Font(size=24, italic=True) # Create a font.
# sheet['A1'].font = italic24Font # Apply the font to A.
# sheet['A1'] = 'Hello, world!'
# wb.save('styles.xlsx')

######################################################################

# Font Objects

# Keyword argument      Data type           Description
# name                  String              The font name, such as 'Calibri' or 'Times New Roman'
# size                  Integer             The point size
# bold                  Boolean             True, for bold font
# italic                Boolean             True, for italic font

# from openpyxl.styles import Font
# wb = openpyxl.Workbook()
# sheet = wb['Sheet']
# fontObj1 = Font(name='Times New Roman', bold=True)
# sheet['A1'].font = fontObj1
# sheet['A1'] = 'Bold Times New Roman'
# fontObj2 = Font(size=24, italic=True)
# sheet['B3'].font = fontObj2
# sheet['B3'] = '24 pt Italic'
# wb.save('styles.xlsx')

######################################################################

# Formulas

# sheet['B9'] = '=SUM(B1:B8)'
# wb = openpyxl.Workbook()
# sheet = wb.active
# sheet['A1'] = 200
# sheet['A2'] = 300
# sheet['A3'] = '=SUM(A1:A2)' # Set the formula.
# wb.save('writeFormula.xlsx')

######################################################################

# Setting Row Height and Column Width

# wb = openpyxl.Workbook()
# sheet = wb.active
# sheet['A1'] = 'Tall row'
# sheet['B2'] = 'Wide column'
# # Set the height and width:
# sheet.row_dimentions[1].height = 70
# sheet.column_dimensions['B'].width = 20
# wb.save('dimensions.xlsx')

######################################################################

# Merging and Unmerging Cells

# wb = openpyxl.Workbook()
# sheet = wb.active
# sheet.merge_cells('A1:D3') # Merge all these cells.
# sheet['A1'] = 'Twelve cells merged together.'
# sheet.merge_cells('C5:D5') # Merge these two cells.
# sheet['C5'] = 'Two merged cells'
# wb.save('merged.xlsx')

# wb = openpyxl.load_workbook('merged.xlsx')
# sheet = wb.active
# sheet.unmerge_cells('A1:D3') # Split these cells up.
# sheet.unmerge_cells('C5:D5')
# wb.save('merged.xlsx')

######################################################################

# Freezing Panes

# freeze_panes setting                    Rows and columns frozen
# sheet.freeze_panes = 'A2'               Row 1
# sheet.freeze_panes = 'B1'               Column A
# sheet.freeze_panes = 'C1'               Columns A and B
# sheet.freeze_panes = 'C2'               Row 1 and columns A and B
# sheet.freeze_panes =                    No frozen panes
# 'A1' or sheet.freeze_panes = None

# wb = openpyxl.load_workbook('produceSales.xlsx')
# sheet = wb.active
# sheet.freeze_panes = 'A2' # Freeze the rows above A2
# wb.save('freezeExample.xlsx')

######################################################################

# Charts

# wb = openpyxl.Workbook()
# sheet = wb.active
# for i in range(1, 11): # Create some data in column A
#     sheet['A' + str(i)] = i

# refObj = openpyxl.chart.Reference(sheet, min_col=1, min_row=1, max_col=1, max_row=10)
# seriesObj = openpyxl.chart.Series(refObj, title='First Series')
# chartObj = openpyxl.chart.BarChart()
# chartObj.title = 'My Chart'
# chartObj.appen(seriesObj)
# sheet.add_chart(chartObj, 'C5')
# wb.save('sampleChart.xlsx')