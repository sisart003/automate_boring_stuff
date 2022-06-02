# Your program does the following:
# 1. Loops over all the rows
# 2. If the row is for garlic, celery, or lemons, changes the price

# This means your code will need to do the following:

# 1. Open the spreadsheet file.
# 2. For each row, check whether the value in column A is Celery, Garlic, or Lemon.
# 3. If it is, update the price in column B.
# 4. Save the spreadsheet to a new file (so that you don’t lose the old spreadsheet, just in case).

#! python3
# Corrects costs in produce sales spreadsheet.

import openpyxl

wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb['Sheet']

# The produce types and their updated prices
PRICE_UPDATES = {'Garlic' : 3.07,
                'Celery' : 1.19,
                'Lemon' : 1.27}

# Loop through the rows and update the prices.
for rowNum in range(2, sheet.max_row): # skip the first row
    produceName = sheet.cell(row=rowNum, column=1).value
    if produceName in PRICE_UPDATES:
        sheet.cell(row=rowNum, column=2).value = PRICE_UPDATES[produceName]

wb.save('updatedProduceSales.xlsx')

# Ideas for Similar Programs
# Since many office workers use Excel spreadsheets all the time, a program
# that can automatically edit and write Excel files could be really useful. Such
# a program could do the following:
# •	 Read data from one spreadsheet and write it to parts of other spreadsheets.
# •	 Read data from websites, text files, or the clipboard and write it to aspreadsheet.
# •	 Automatically “clean up” data in spreadsheets. For example, it could use regular expressions to read multiple formats of phone numbers and edit them to a single, standard format.